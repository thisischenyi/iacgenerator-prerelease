"""Test Excel parsing with metadata to tags merge."""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.services.excel_parser import ExcelParserService
from openpyxl import Workbook, load_workbook
from io import BytesIO

# Create a test Excel file
wb = Workbook()
ws = wb.active
ws.title = "Azure_VM"

# Add headers (matching the user's Excel structure)
headers = [
    "ResourceName*",
    "Environment*",
    "Project*",
    "Owner",
    "CostCenter",
    "Tags",
    "ResourceGroup*",
    "Location*",
    "VMSize*",
    "OSType*",
    "ImagePublisher",
    "ImageOffer",
    "ImageSKU",
    "ImageVersion",
]
for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=header)

# Add data row (matching user's example)
data_row = [
    "web-vm-01",  # ResourceName
    "Production",  # Environment
    "abc",  # Project
    "john.doe@example.com",  # Owner
    "IT-1234",  # CostCenter
    '{"Application": "WebServer"}',  # Tags (JSON)
    "rg-myproject-prod",  # ResourceGroup
    "eastus",  # Location
    "Standard_D2s_v3",  # VMSize
    "Linux",  # OSType
    "Canonical",  # ImagePublisher
    "UbuntuServer",  # ImageOffer
    "18.04-LTS",  # ImageSKU
    "latest",  # ImageVersion
]
for col_idx, value in enumerate(data_row, start=1):
    ws.cell(row=2, column=col_idx, value=value)

# Save to BytesIO
excel_bytes = BytesIO()
wb.save(excel_bytes)
excel_bytes.seek(0)

# Parse the Excel
parser = ExcelParserService()
result = parser.parse_excel_file(excel_bytes.read())

print("=" * 80)
print("EXCEL PARSING TEST")
print("=" * 80)
print(f"Success: {result.success}")
print(f"Resource Count: {result.resource_count}")
print(f"Resource Types: {result.resource_types}")
print(f"Errors: {result.errors}")
print(f"Warnings: {result.warnings}")
print()

# Debug: Print workbook info
excel_bytes.seek(0)
wb_debug = load_workbook(BytesIO(excel_bytes.read()))
print(f"DEBUG: Sheet names: {wb_debug.sheetnames}")
ws_debug = wb_debug["Azure_VM"]
print(f"DEBUG: Max row: {ws_debug.max_row}")
print(f"DEBUG: Max col: {ws_debug.max_column}")
print(f"DEBUG: Row 1 (headers):")
for cell in ws_debug[1]:
    print(f"  {cell.value}")
print(f"DEBUG: Row 2 (data):")
for cell in ws_debug[2]:
    print(f"  {cell.value}")
print()

if result.resources:
    resource = result.resources[0]
    print(f"Resource Name: {resource.resource_name}")
    print(f"Resource Type: {resource.resource_type}")
    print(f"Cloud Platform: {resource.cloud_platform}")
    print()
    print("Properties (top-level):")
    print(f"  Environment: {resource.properties.get('Environment')}")
    print(f"  Project: {resource.properties.get('Project')}")
    print(f"  Owner: {resource.properties.get('Owner')}")
    print(f"  CostCenter: {resource.properties.get('CostCenter')}")
    print()
    print("Tags dict:")
    tags = resource.properties.get("Tags", {})
    for key, value in tags.items():
        print(f"  {key}: {value}")
    print()

    # Verify metadata was merged into Tags
    if isinstance(tags, dict):
        has_environment = any(k.lower() == "environment" for k in tags.keys())
        has_project = any(k.lower() == "project" for k in tags.keys())
        has_owner = any(k.lower() == "owner" for k in tags.keys())
        has_costcenter = any(k.lower() == "costcenter" for k in tags.keys())
        has_application = any(k.lower() == "application" for k in tags.keys())

        print("Verification:")
        print(f"  Environment in Tags: {has_environment} (expected: True)")
        print(f"  Project in Tags: {has_project} (expected: True)")
        print(f"  Owner in Tags: {has_owner} (expected: True)")
        print(f"  CostCenter in Tags: {has_costcenter} (expected: True)")
        print(f"  Application in Tags: {has_application} (expected: True, from JSON)")
        print()

        if (
            has_environment
            and has_project
            and has_owner
            and has_costcenter
            and has_application
        ):
            print("[SUCCESS] Metadata successfully merged into Tags!")
            print("Compliance checker will now be able to validate Project tag.")
        else:
            print("[FAILED] Metadata merge incomplete!")
            sys.exit(1)
    else:
        print(f"[ERROR] Tags is not a dict: {type(tags)}")
        sys.exit(1)
else:
    print("[ERROR] No resources parsed!")
    sys.exit(1)
