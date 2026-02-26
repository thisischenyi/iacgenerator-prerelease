"""Excel parsing service for resource definitions."""

import io
import re
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas import (
    ResourceInfo,
    ExcelParseResult,
    CloudPlatform,
)


class ExcelParserService:
    """Service for parsing Excel files containing resource definitions."""

    # Supported resource types for each cloud platform
    AWS_RESOURCE_TYPES = [
        "AWS_EC2",
        "AWS_VPC",
        "AWS_Subnet",
        "AWS_SecurityGroup",
        "AWS_S3",
        "AWS_RDS",
        "AWS_InternetGateway",
        "AWS_NATGateway",
        "AWS_ElasticIP",
        "AWS_LoadBalancer",
        "AWS_TargetGroup",
    ]

    AZURE_RESOURCE_TYPES = [
        "Azure_VM",
        "Azure_VNet",
        "Azure_Subnet",
        "Azure_NSG",
        "Azure_Storage",
        "Azure_SQL",
        "Azure_PublicIP",
        "Azure_NATGateway",
        "Azure_LoadBalancer",
    ]

    # Common fields for all resources
    COMMON_FIELDS = [
        "ResourceName",
        "Environment",
        "Project",
        "Owner",
        "CostCenter",
        "Tags",
    ]

    def __init__(self):
        """Initialize Excel parser service."""
        self.errors: List[str] = []
        self.warnings: List[str] = []

    def parse_excel_file(self, file_content: bytes) -> ExcelParseResult:
        """
        Parse Excel file and extract resource definitions.

        Args:
            file_content: Excel file content as bytes

        Returns:
            ExcelParseResult with parsed resources and any errors/warnings
        """
        self.errors = []
        self.warnings = []

        try:
            # Load workbook from bytes
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)

            # Parse all resource sheets
            resources: List[ResourceInfo] = []
            resource_types: List[str] = []

            for sheet_name in workbook.sheetnames:
                # Skip README and other non-resource sheets
                if sheet_name.upper() == "README" or not self._is_resource_sheet(
                    sheet_name
                ):
                    continue

                sheet = workbook[sheet_name]
                cloud_platform = self._get_cloud_platform(sheet_name)

                if not cloud_platform:
                    self.warnings.append(f"Skipping unknown sheet: {sheet_name}")
                    continue

                # Parse resources from this sheet
                sheet_resources = self._parse_resource_sheet(
                    sheet, sheet_name, cloud_platform
                )

                if sheet_resources:
                    resources.extend(sheet_resources)
                    if sheet_name not in resource_types:
                        resource_types.append(sheet_name)

            return ExcelParseResult(
                success=len(self.errors) == 0,
                resource_count=len(resources),
                resource_types=resource_types,
                resources=resources,
                errors=self.errors if self.errors else None,
                warnings=self.warnings if self.warnings else None,
            )

        except Exception as e:
            self.errors.append(f"Failed to parse Excel file: {str(e)}")
            return ExcelParseResult(
                success=False,
                resource_count=0,
                resource_types=[],
                resources=[],
                errors=self.errors,
                warnings=None,
            )

    def _is_resource_sheet(self, sheet_name: str) -> bool:
        """
        Check if sheet name represents a resource type.

        Args:
            sheet_name: Name of the sheet

        Returns:
            True if sheet is a resource sheet
        """
        return (
            sheet_name in self.AWS_RESOURCE_TYPES
            or sheet_name in self.AZURE_RESOURCE_TYPES
        )

    def _get_cloud_platform(self, sheet_name: str) -> Optional[CloudPlatform]:
        """
        Determine cloud platform from sheet name.

        Args:
            sheet_name: Name of the sheet

        Returns:
            CloudPlatform enum or None
        """
        if sheet_name.startswith("AWS_"):
            return CloudPlatform.AWS
        elif sheet_name.startswith("Azure_"):
            return CloudPlatform.AZURE
        return None

    def _parse_resource_sheet(
        self,
        sheet: Worksheet,
        sheet_name: str,
        cloud_platform: CloudPlatform,
    ) -> List[ResourceInfo]:
        """
        Parse a single resource sheet.

        Args:
            sheet: Excel worksheet
            sheet_name: Name of the sheet
            cloud_platform: Cloud platform (AWS or Azure)

        Returns:
            List of parsed resources
        """
        resources: List[ResourceInfo] = []

        # Get header row (first row)
        headers = []
        for cell in sheet[1]:
            if cell.value:
                # Strip asterisks from required field markers
                header = str(cell.value).strip().rstrip("*")
                headers.append(header)
            else:
                headers.append("")

        if not headers:
            self.errors.append(f"Sheet {sheet_name}: No headers found")
            return resources

        # Parse data rows (starting from row 2)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            # Skip empty rows
            if all(cell.value is None for cell in row):
                continue

            # Build resource properties from row
            properties: Dict[str, Any] = {}
            resource_name: Optional[str] = None

            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    break

                header = headers[col_idx]
                if not header:
                    continue

                value = cell.value

                # Store ResourceName separately
                if header == "ResourceName":
                    resource_name = str(value) if value else None

                # Store all properties
                if value is not None:
                    converted_val = self._convert_cell_value(value, header)
                    # For list-like fields, ensure they are lists
                    list_headers = [
                        "Subnets",
                        "SecurityGroups",
                        "SecurityGroupIds",
                        "AddressSpace",
                        "DnsServers",
                        "ServiceEndpoints",
                        "BlobContainers",
                        "Targets",
                        "BackendPoolResources",
                    ]
                    if header in list_headers and isinstance(converted_val, str):
                        # Convert comma-separated string to list
                        if "," in converted_val:
                            converted_val = [
                                x.strip() for x in converted_val.split(",") if x.strip()
                            ]
                        else:
                            converted_val = [converted_val]

                    properties[header] = converted_val

            # Validate and create resource
            if resource_name:
                # Extract resource type from sheet name
                resource_type = (
                    sheet_name.split("_", 1)[1] if "_" in sheet_name else sheet_name
                )

                # Inject secure defaults
                if resource_type == "S3":
                    if "PublicAccess" not in properties:
                        properties["PublicAccess"] = False
                elif resource_type == "Storage":
                    if "EnableHttpsTrafficOnly" not in properties:
                        properties["EnableHttpsTrafficOnly"] = True
                    if "MinTlsVersion" not in properties:
                        properties["MinTlsVersion"] = "TLS1_2"
                elif resource_type == "Subnet" and cloud_platform == CloudPlatform.AZURE:
                    # Guardrail: strip SQL-style endpoint values that commonly lead to invalid subnet configs.
                    self._sanitize_azure_subnet_service_endpoints(
                        properties=properties,
                        sheet_name=sheet_name,
                        row_idx=row_idx,
                        resource_name=resource_name,
                    )
                elif resource_type == "SQL" and cloud_platform == CloudPlatform.AZURE:
                    # Guardrail: prevent incompatible SQL networking/auditing combinations.
                    self._normalize_azure_sql_properties(
                        properties=properties,
                        sheet_name=sheet_name,
                        row_idx=row_idx,
                        resource_name=resource_name,
                    )
                elif resource_type == "LoadBalancer" and cloud_platform == CloudPlatform.AZURE:
                    # Normalize Azure LoadBalancer field aliases so template rendering stays consistent.
                    self._normalize_azure_load_balancer_properties(properties)

                # Merge metadata columns (Environment, Project, Owner, CostCenter) into Tags
                # This ensures compliance checks can validate these as tags
                self._merge_metadata_to_tags(properties)

                resource = ResourceInfo(
                    resource_type=resource_type,
                    cloud_platform=cloud_platform,
                    resource_name=resource_name,
                    properties=properties,
                )
                resources.append(resource)
            else:
                self.warnings.append(
                    f"Sheet {sheet_name}, Row {row_idx}: Missing ResourceName, skipping row"
                )

        return resources

    def _convert_cell_value(self, value: Any, header: str) -> Any:
        """
        Convert cell value to appropriate Python type.

        Args:
            value: Cell value
            header: Column header name

        Returns:
            Converted value
        """
        # Handle None
        if value is None:
            return None

        # Convert to string first
        str_value = str(value).strip()

        # Handle JSON fields (Tags, IngressRules, etc.)
        if header in [
            "Tags",
            "IngressRules",
            "EgressRules",
            "SecurityRules",
            "DataDisks",
            "LifecycleRules",
            "BlobContainers",
            "NetworkRules",
            "FirewallRules",
            "VirtualNetworkRules",
            "LongTermRetention",
        ]:
            # Try to parse as JSON
            import json

            try:
                return json.loads(str_value)
            except json.JSONDecodeError:
                self.warnings.append(
                    f"Invalid JSON in field {header}: {str_value[:50]}..."
                )
                return str_value

        # Handle boolean fields
        if isinstance(value, bool):
            return value
        if str_value.lower() in ["true", "yes", "1"]:
            return True
        if str_value.lower() in ["false", "no", "0"]:
            return False

        # Handle numeric fields
        if isinstance(value, (int, float)):
            return value

        # Try to convert to number
        try:
            if "." in str_value:
                return float(str_value)
            else:
                return int(str_value)
        except ValueError:
            pass

        # Return as string
        return str_value

    def validate_resource(self, resource: ResourceInfo) -> Tuple[bool, List[str]]:
        """
        Validate a single resource definition.

        Args:
            resource: Resource to validate

        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors: List[str] = []

        # Check required common fields
        required_common = ["ResourceName", "Environment", "Project"]
        for field in required_common:
            if field not in resource.properties or not resource.properties[field]:
                errors.append(f"Missing required field: {field}")

        # Validate based on resource type and cloud platform
        if resource.cloud_platform == CloudPlatform.AWS:
            errors.extend(self._validate_aws_resource(resource))
        elif resource.cloud_platform == CloudPlatform.AZURE:
            errors.extend(self._validate_azure_resource(resource))

        return len(errors) == 0, errors

    def _normalize_azure_load_balancer_properties(
        self, properties: Dict[str, Any]
    ) -> None:
        """Normalize Azure LoadBalancer property aliases into canonical keys."""
        aliases = {
            "PrivateIPAllocation": "PrivateIPAddressAllocation",
            "HealthProbeNumberOfProbes": "HealthProbeThreshold",
            "EnableFloatingIP": "LBRuleEnableFloatingIP",
            "IdleTimeoutMinutes": "LBRuleIdleTimeout",
            "DisableOutboundSnat": "LBRuleDisableOutboundSnat",
        }
        for alias, canonical in aliases.items():
            if canonical not in properties and alias in properties:
                properties[canonical] = properties[alias]

        # Canonicalize LB rule protocol to avoid invalid Terraform values.
        lb_rule_protocol = properties.get("LBRuleProtocol")
        if isinstance(lb_rule_protocol, str):
            protocol_lower = lb_rule_protocol.strip().lower()
            protocol_map = {
                "all": "All",
                "tcp": "Tcp",
                "udp": "Udp",
                # Azure LB rule only supports L4 protocols; map common L7 input to Tcp.
                "http": "Tcp",
                "https": "Tcp",
            }
            mapped = protocol_map.get(protocol_lower)
            if mapped:
                if mapped != lb_rule_protocol:
                    properties["LBRuleProtocol"] = mapped
                if protocol_lower in ("http", "https"):
                    self.warnings.append(
                        "Azure LoadBalancer LBRuleProtocol only supports Tcp/Udp/All; "
                        f"mapped '{lb_rule_protocol}' to 'Tcp'."
                    )

    def _sanitize_azure_subnet_service_endpoints(
        self,
        properties: Dict[str, Any],
        sheet_name: str,
        row_idx: int,
        resource_name: str,
    ) -> None:
        """
        Normalize SQL-like ServiceEndpoints values for Azure subnet and emit warning.

        Keep canonical `Microsoft.Sql`. If users input invalid SQL-like values such as
        `Microsoft.Sql/servers`, map them to `Microsoft.Sql` to avoid failed deployments.
        """
        raw_endpoints = properties.get("ServiceEndpoints")
        if raw_endpoints in (None, ""):
            return

        if isinstance(raw_endpoints, str):
            endpoints = [item.strip() for item in raw_endpoints.split(",") if item.strip()]
        elif isinstance(raw_endpoints, list):
            endpoints = [str(item).strip() for item in raw_endpoints if str(item).strip()]
        else:
            endpoints = [str(raw_endpoints).strip()]

        removed: List[str] = []
        mapped_to_sql = False
        kept: List[str] = []
        for endpoint in endpoints:
            normalized = endpoint.lower()
            if normalized == "microsoft.sql":
                kept.append("Microsoft.Sql")
            elif normalized.startswith("microsoft.sql/"):
                removed.append(endpoint)
                mapped_to_sql = True
            else:
                kept.append(endpoint)

        if removed:
            self.warnings.append(
                f"Sheet {sheet_name}, Row {row_idx}, Resource {resource_name}: "
                f"Removed unsupported SQL ServiceEndpoints values {removed}."
            )

        if mapped_to_sql and "Microsoft.Sql" not in kept:
            kept.append("Microsoft.Sql")
            self.warnings.append(
                f"Sheet {sheet_name}, Row {row_idx}, Resource {resource_name}: "
                "Mapped invalid SQL ServiceEndpoints values to 'Microsoft.Sql'."
            )

        if kept:
            properties["ServiceEndpoints"] = kept
        else:
            properties.pop("ServiceEndpoints", None)

    def _normalize_azure_sql_properties(
        self,
        properties: Dict[str, Any],
        sheet_name: str,
        row_idx: int,
        resource_name: str,
    ) -> None:
        """Normalize Azure SQL properties to avoid known deployment conflicts."""

        public_network_access = str(
            properties.get("PublicNetworkAccess", "true")
        ).strip().lower()
        public_network_disabled = public_network_access in {
            "false",
            "disabled",
            "deny",
            "no",
            "0",
        }

        if public_network_disabled:
            if properties.get("FirewallRules"):
                properties.pop("FirewallRules", None)
                self.warnings.append(
                    f"Sheet {sheet_name}, Row {row_idx}, Resource {resource_name}: "
                    "Removed FirewallRules because PublicNetworkAccess is disabled."
                )
            if properties.get("VirtualNetworkRules"):
                properties.pop("VirtualNetworkRules", None)
                self.warnings.append(
                    f"Sheet {sheet_name}, Row {row_idx}, Resource {resource_name}: "
                    "Removed VirtualNetworkRules because PublicNetworkAccess is disabled."
                )

        auditing_enabled = str(properties.get("AuditingEnabled", "false")).strip().lower()
        if auditing_enabled in {"true", "enabled", "yes", "1"}:
            endpoint = str(properties.get("AuditingStorageEndpoint", "")).strip()
            is_blob_endpoint = (
                endpoint.startswith("https://") and ".blob.core.windows.net" in endpoint
            )
            if not is_blob_endpoint:
                properties["AuditingEnabled"] = "false"
                self.warnings.append(
                    f"Sheet {sheet_name}, Row {row_idx}, Resource {resource_name}: "
                    "Disabled AuditingEnabled because AuditingStorageEndpoint is missing "
                    "or invalid (expected blob endpoint)."
                )

    def _validate_aws_resource(self, resource: ResourceInfo) -> List[str]:
        """Validate AWS-specific resource fields."""
        errors: List[str] = []
        props = resource.properties

        if resource.resource_type == "EC2":
            required = [
                "Region",
                "InstanceType",
                "AMI_ID",
                "VPC",
                "VPCExists",  # New required field
                "Subnet",
                "SubnetExists",  # New required field
                "SecurityGroups",
                "SecurityGroupsExist",  # New required field
                "KeyPairName",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for EC2: {field}")
            # Validate Exists fields values
            exists_fields = ["VPCExists", "SubnetExists", "SecurityGroupsExist"]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")

        elif resource.resource_type == "VPC":
            required = ["Region", "CIDR_Block"]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for VPC: {field}")

            # Validate CIDR format
            if "CIDR_Block" in props:
                if not self._is_valid_cidr(props["CIDR_Block"]):
                    errors.append(f"Invalid CIDR format: {props['CIDR_Block']}")

        elif resource.resource_type == "Subnet":
            required = [
                "VPC",
                "VPCExists",
                "AvailabilityZone",
                "CIDR_Block",
            ]  # Updated required fields
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for Subnet: {field}")
            # Validate VPCExists value
            if "VPCExists" in props and props["VPCExists"] not in ["y", "n"]:
                errors.append("VPCExists must be 'y' or 'n'")

        elif resource.resource_type == "SecurityGroup":
            required = [
                "VPC",
                "VPCExists",
                "Description",
                "IngressRules",
            ]  # Updated required fields
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for SecurityGroup: {field}")
            # Validate VPCExists value
            if "VPCExists" in props and props["VPCExists"] not in ["y", "n"]:
                errors.append("VPCExists must be 'y' or 'n'")

        elif resource.resource_type == "S3":
            required = ["Region", "Versioning", "Encryption"]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for S3: {field}")

        elif resource.resource_type == "RDS":
            required = [
                "Region",
                "Engine",
                "InstanceClass",
                "AllocatedStorage",
                "DBName",
                "MasterUsername",
                "VPC",
                "VPCExists",  # New required field
                "SecurityGroups",
                "SecurityGroupsExist",  # New required field
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for RDS: {field}")
            # Validate Exists fields values
            exists_fields = ["VPCExists", "SecurityGroupsExist"]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")

        elif resource.resource_type == "InternetGateway":
            required = [
                "Region",
                "VPC",
                "VPCExists",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(
                        f"Missing required field for InternetGateway: {field}"
                    )
            # Validate VPCExists value
            if "VPCExists" in props and props["VPCExists"] not in ["y", "n"]:
                errors.append("VPCExists must be 'y' or 'n'")

        elif resource.resource_type == "NATGateway":
            required = [
                "Region",
                "Subnet",
                "SubnetExists",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for NATGateway: {field}")
            # Validate Exists fields values
            exists_fields = ["SubnetExists", "InternetGatewayExists"]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")
            # Validate ConnectivityType if provided
            if "ConnectivityType" in props:
                if props["ConnectivityType"] not in ["public", "private"]:
                    errors.append("ConnectivityType must be 'public' or 'private'")

        elif resource.resource_type == "ElasticIP":
            required = [
                "Region",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for ElasticIP: {field}")
            # Validate Exists fields values
            exists_fields = ["InstanceExists", "NetworkInterfaceExists"]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")
            # Validate Domain if provided
            if "Domain" in props:
                if props["Domain"] not in ["vpc", "standard"]:
                    errors.append("Domain must be 'vpc' or 'standard'")

        elif resource.resource_type == "LoadBalancer":
            required = [
                "Region",
                "Type",
                "Scheme",
                "VPC",
                "VPCExists",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for LoadBalancer: {field}")
            # Validate Exists fields values
            exists_fields = [
                "VPCExists",
                "SubnetExists",
                "SecurityGroupsExist",
                "ListenerTargetGroupExists",
            ]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")
            # Validate Type
            if "Type" in props:
                if props["Type"] not in ["application", "network"]:
                    errors.append("Type must be 'application' or 'network'")
            # Validate Scheme
            if "Scheme" in props:
                if props["Scheme"] not in ["internet-facing", "internal"]:
                    errors.append("Scheme must be 'internet-facing' or 'internal'")
            # Validate IPAddressType if provided
            if "IPAddressType" in props:
                if props["IPAddressType"] not in ["ipv4", "dualstack"]:
                    errors.append("IPAddressType must be 'ipv4' or 'dualstack'")
            # Validate IdleTimeout for ALB
            if "IdleTimeout" in props and "Type" in props:
                if props["Type"] == "application":
                    try:
                        timeout = int(props["IdleTimeout"])
                        if timeout < 1 or timeout > 4000:
                            errors.append(
                                "IdleTimeout must be between 1 and 4000 seconds for ALB"
                            )
                    except (ValueError, TypeError):
                        errors.append("IdleTimeout must be a valid integer")
            # Validate ListenerProtocol if provided
            if "ListenerProtocol" in props:
                if props["ListenerProtocol"] not in [
                    "HTTP",
                    "HTTPS",
                    "TCP",
                    "UDP",
                    "TLS",
                ]:
                    errors.append(
                        "ListenerProtocol must be 'HTTP', 'HTTPS', 'TCP', 'UDP', or 'TLS'"
                    )
            # Validate boolean fields
            bool_fields = ["CrossZoneEnabled", "DeletionProtection"]
            for field in bool_fields:
                if field in props:
                    val = str(props[field]).lower()
                    if val not in ["true", "false"]:
                        errors.append(f"{field} must be 'true' or 'false'")

        elif resource.resource_type == "TargetGroup":
            required = [
                "Region",
                "Port",
                "Protocol",
                "VPC",
                "VPCExists",
                "TargetType",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for TargetGroup: {field}")
            # Validate VPCExists
            if "VPCExists" in props and props["VPCExists"] not in ["y", "n"]:
                errors.append("VPCExists must be 'y' or 'n'")
            # Validate Protocol
            if "Protocol" in props:
                if props["Protocol"] not in [
                    "HTTP",
                    "HTTPS",
                    "TCP",
                    "UDP",
                    "TLS",
                    "GENEVE",
                ]:
                    errors.append(
                        "Protocol must be 'HTTP', 'HTTPS', 'TCP', 'UDP', 'TLS', or 'GENEVE'"
                    )
            # Validate TargetType
            if "TargetType" in props:
                if props["TargetType"] not in ["instance", "ip", "lambda", "alb"]:
                    errors.append(
                        "TargetType must be 'instance', 'ip', 'lambda', or 'alb'"
                    )
            # Validate Port
            if "Port" in props:
                try:
                    port = int(props["Port"])
                    if port < 1 or port > 65535:
                        errors.append("Port must be between 1 and 65535")
                except (ValueError, TypeError):
                    errors.append("Port must be a valid integer between 1 and 65535")
            # Validate HealthCheckProtocol if provided
            if "HealthCheckProtocol" in props:
                if props["HealthCheckProtocol"] not in ["HTTP", "HTTPS", "TCP"]:
                    errors.append(
                        "HealthCheckProtocol must be 'HTTP', 'HTTPS', or 'TCP'"
                    )
            # Validate numeric health check fields
            numeric_fields = {
                "HealthCheckInterval": (5, 300),
                "HealthyThreshold": (2, 10),
                "UnhealthyThreshold": (2, 10),
                "HealthCheckTimeout": (2, 120),
                "DeregistrationDelay": (0, 3600),
                "SlowStart": (30, 900),
            }
            for field, (min_val, max_val) in numeric_fields.items():
                if field in props:
                    try:
                        val = int(props[field])
                        if val < min_val or val > max_val:
                            errors.append(
                                f"{field} must be between {min_val} and {max_val}"
                            )
                    except (ValueError, TypeError):
                        errors.append(f"{field} must be a valid integer")
            # Validate boolean fields
            bool_fields = ["StickinessEnabled"]
            for field in bool_fields:
                if field in props:
                    val = str(props[field]).lower()
                    if val not in ["true", "false"]:
                        errors.append(f"{field} must be 'true' or 'false'")

        return errors

    def _validate_azure_resource(self, resource: ResourceInfo) -> List[str]:
        """Validate Azure-specific resource fields."""
        errors: List[str] = []
        props = resource.properties

        if resource.resource_type == "VM":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",  # New required field
                "VNet",
                "VNetExists",  # New required field
                "Subnet",
                "SubnetExists",  # New required field
                "NSG",
                "NSGExists",  # New required field
                "Location",
                "VMSize",
                "OSType",
                "AdminUsername",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for VM: {field}")
            # Validate Exists fields values
            exists_fields = [
                "ResourceGroupExists",
                "VNetExists",
                "SubnetExists",
                "NSGExists",
            ]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")

        elif resource.resource_type == "VNet":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",
                "Location",
                "AddressSpace",
            ]  # Updated required fields
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for VNet: {field}")
            # Validate ResourceGroupExists value
            if "ResourceGroupExists" in props and props["ResourceGroupExists"] not in [
                "y",
                "n",
            ]:
                errors.append("ResourceGroupExists must be 'y' or 'n'")

        elif resource.resource_type == "Subnet":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",
                "VNet",
                "VNetExists",
                "AddressPrefix",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for Subnet: {field}")
            # Validate Exists field values
            exists_fields = ["ResourceGroupExists", "VNetExists"]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")

        elif resource.resource_type == "NSG":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",
                "Location",
                "SecurityRules",
            ]  # Updated required fields
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for NSG: {field}")
            # Validate ResourceGroupExists value
            if "ResourceGroupExists" in props and props["ResourceGroupExists"] not in [
                "y",
                "n",
            ]:
                errors.append("ResourceGroupExists must be 'y' or 'n'")

        elif resource.resource_type == "Storage":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",  # New required field
                "Location",
                "AccountKind",
                "AccountTier",
                "ReplicationType",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for Storage: {field}")
            # Validate ResourceGroupExists value
            if "ResourceGroupExists" in props and props["ResourceGroupExists"] not in [
                "y",
                "n",
            ]:
                errors.append("ResourceGroupExists must be 'y' or 'n'")

        elif resource.resource_type == "SQL":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",  # New required field
                "Location",
                "ServerName",
                "ServerAdminLogin",
                "DatabaseEdition",
                "VNet",
                "VNetExists",  # New required field
                "Subnet",
                "SubnetExists",  # New required field
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for SQL: {field}")
            # Validate Exists fields values
            exists_fields = ["ResourceGroupExists", "VNetExists", "SubnetExists"]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")

        elif resource.resource_type == "PublicIP":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",
                "Location",
                "AllocationMethod",
                "SKU",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for PublicIP: {field}")
            # Validate ResourceGroupExists value
            if "ResourceGroupExists" in props and props["ResourceGroupExists"] not in [
                "y",
                "n",
            ]:
                errors.append("ResourceGroupExists must be 'y' or 'n'")
            # Validate AllocationMethod
            if "AllocationMethod" in props:
                if props["AllocationMethod"] not in ["Static", "Dynamic"]:
                    errors.append("AllocationMethod must be 'Static' or 'Dynamic'")
            # Validate SKU
            if "SKU" in props:
                if props["SKU"] not in ["Basic", "Standard"]:
                    errors.append("SKU must be 'Basic' or 'Standard'")

        elif resource.resource_type == "NATGateway":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",
                "Location",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for NATGateway: {field}")
            # Validate Exists fields values
            exists_fields = ["ResourceGroupExists", "PublicIPExists", "SubnetExists"]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")
            # Validate IdleTimeoutMinutes if provided
            if "IdleTimeoutMinutes" in props:
                timeout = props["IdleTimeoutMinutes"]
                if isinstance(timeout, (int, float)):
                    if timeout < 4 or timeout > 120:
                        errors.append("IdleTimeoutMinutes must be between 4 and 120")

        elif resource.resource_type == "LoadBalancer":
            required = [
                "ResourceGroup",
                "ResourceGroupExists",
                "Location",
                "SKU",
                "FrontendIPName",
            ]
            for field in required:
                if field not in props or not props[field]:
                    errors.append(f"Missing required field for LoadBalancer: {field}")
            # Validate Exists fields values
            exists_fields = ["ResourceGroupExists", "PublicIPExists", "SubnetExists"]
            for field in exists_fields:
                if field in props and props[field] not in ["y", "n"]:
                    errors.append(f"{field} must be 'y' or 'n'")
            # Validate SKU
            if "SKU" in props:
                if props["SKU"] not in ["Basic", "Standard"]:
                    errors.append("SKU must be 'Basic' or 'Standard'")
            # Validate HealthProbeProtocol if provided
            if "HealthProbeProtocol" in props:
                if props["HealthProbeProtocol"] not in ["Tcp", "Http", "Https"]:
                    errors.append(
                        "HealthProbeProtocol must be 'Tcp', 'Http', or 'Https'"
                    )
            # Validate LBRuleProtocol if provided
            if "LBRuleProtocol" in props:
                if props["LBRuleProtocol"] not in ["Tcp", "Udp", "All"]:
                    errors.append("LBRuleProtocol must be 'Tcp', 'Udp', or 'All'")
            # Validate BackendPoolResources if provided
            if "BackendPoolResources" in props:
                backend_resources = props["BackendPoolResources"]
                if not isinstance(backend_resources, list):
                    errors.append("BackendPoolResources must be a comma-separated list")
                elif not all(
                    isinstance(resource_name, str) and resource_name.strip()
                    for resource_name in backend_resources
                ):
                    errors.append(
                        "BackendPoolResources entries must be non-empty resource names"
                    )

        return errors

    def _is_valid_cidr(self, cidr: str) -> bool:
        """
        Validate CIDR notation.

        Args:
            cidr: CIDR string to validate

        Returns:
            True if valid CIDR
        """
        pattern = r"^([0-9]{1,3}\.){3}[0-9]{1,3}/([0-9]|[1-2][0-9]|3[0-2])$"
        return bool(re.match(pattern, str(cidr)))

    def _merge_metadata_to_tags(self, properties: Dict[str, Any]) -> None:
        """
        Merge metadata columns (Environment, Project, Owner, CostCenter) into Tags dict.

        This ensures that these common metadata fields are available as tags
        for compliance checking and Terraform resource tagging.

        Args:
            properties: Resource properties dict (modified in-place)
        """
        # Get existing Tags dict or create new one
        tags = properties.get("Tags", {})
        if not isinstance(tags, dict):
            # If Tags is not a dict (e.g., parsing failed), create empty dict
            tags = {}

        # Metadata fields to merge into tags
        metadata_fields = ["Environment", "Project", "Owner", "CostCenter"]

        for field in metadata_fields:
            if field in properties and properties[field]:
                # Only add to tags if not already present (case-insensitive check)
                tag_keys_lower = {k.lower(): k for k in tags.keys()}
                if field.lower() not in tag_keys_lower:
                    tags[field] = properties[field]

        # Update Tags in properties
        properties["Tags"] = tags
