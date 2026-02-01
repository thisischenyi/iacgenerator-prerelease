"""
Comprehensive test for Excel template improvements.

This test verifies:
1. Required fields are marked with asterisk (*)
2. Required fields have darker purple background (#7D00C7)
3. Optional fields have standard purple background (#A100FF)
4. Sample data is present in row 2 of each sheet
5. Parser correctly handles asterisks in headers
6. README sheet has updated instructions
"""

import io
from openpyxl import load_workbook
from app.services.excel_generator import ExcelGeneratorService
from app.services.excel_parser import ExcelParserService
from app.schemas import TemplateType


def test_required_field_markers():
    """Test that required fields are marked with asterisk and different color."""
    service = ExcelGeneratorService()
    template_bytes = service.generate_template(TemplateType.AWS)
    wb = load_workbook(io.BytesIO(template_bytes))

    sheet = wb["AWS_EC2"]

    # Check ResourceName is marked as required
    resource_name_cell = sheet.cell(1, 1)
    assert "*" in str(resource_name_cell.value), "ResourceName should have asterisk"
    assert resource_name_cell.fill.start_color.rgb == "FF7D00C7", (
        "Required field should have darker purple"
    )

    # Check Owner is optional
    owner_cell = sheet.cell(1, 4)
    assert "*" not in str(owner_cell.value), "Owner should NOT have asterisk"
    assert owner_cell.fill.start_color.rgb == "FFA100FF", (
        "Optional field should have standard purple"
    )

    print("PASS: Required field markers working correctly")


def test_sample_data_row():
    """Test that sample data is present in row 2."""
    service = ExcelGeneratorService()
    template_bytes = service.generate_template(TemplateType.FULL)
    wb = load_workbook(io.BytesIO(template_bytes))

    # Test AWS_EC2
    ec2_sheet = wb["AWS_EC2"]
    assert ec2_sheet.cell(2, 1).value == "web-server-01", (
        "AWS_EC2 sample ResourceName should be 'web-server-01'"
    )
    assert ec2_sheet.cell(2, 2).value == "Production", (
        "AWS_EC2 sample Environment should be 'Production'"
    )
    assert ec2_sheet.cell(2, 9).value == "t3.medium", (
        "AWS_EC2 sample InstanceType should be 't3.medium'"
    )

    # Test Azure_VM
    vm_sheet = wb["Azure_VM"]
    assert vm_sheet.cell(2, 1).value == "web-vm-01", (
        "Azure_VM sample ResourceName should be 'web-vm-01'"
    )
    assert vm_sheet.cell(2, 7).value == "rg-myproject-prod", (
        "Azure_VM sample ResourceGroup should be 'rg-myproject-prod'"
    )
    assert vm_sheet.cell(2, 10).value == "Linux", (
        "Azure_VM sample OSType should be 'Linux'"
    )

    print("PASS: Sample data row working correctly")


def test_parser_strips_asterisks():
    """Test that parser correctly strips asterisks from headers."""
    service = ExcelGeneratorService()
    template_bytes = service.generate_template(TemplateType.AWS)

    parser = ExcelParserService()
    result = parser.parse_excel_file(template_bytes)

    assert result.success, "Parser should succeed"
    assert result.resource_count == 6, (
        f"Should parse 6 sample resources (AWS only), got {result.resource_count}"
    )

    # Check that ResourceName property exists (not ResourceName*)
    ec2_resource = next((r for r in result.resources if r.resource_type == "EC2"), None)
    assert ec2_resource is not None, "Should find EC2 resource"
    assert "ResourceName" in ec2_resource.properties, (
        "Should have 'ResourceName' property (without asterisk)"
    )
    assert ec2_resource.properties["ResourceName"] == "web-server-01", (
        "Should parse sample data correctly"
    )

    print("PASS: Parser strips asterisks correctly")


def test_readme_instructions():
    """Test that README has updated instructions."""
    service = ExcelGeneratorService()
    template_bytes = service.generate_template(TemplateType.FULL)
    wb = load_workbook(io.BytesIO(template_bytes))

    readme = wb["README"]

    # Check for required fields instruction
    required_field_instruction = readme.cell(4, 2).value  # Row 4 is "Required Fields"
    assert "asterisk (*)" in required_field_instruction, (
        "Should mention asterisk in required fields instruction"
    )
    assert "#7D00C7" in required_field_instruction, "Should mention darker purple color"

    # Check for sample data instruction
    sample_data_instruction = readme.cell(5, 2).value  # Row 5 is "Sample Data"
    assert "Row 2" in sample_data_instruction, "Should mention Row 2 for sample data"
    assert "sample data" in sample_data_instruction.lower(), (
        "Should explain sample data"
    )

    print("PASS: README instructions updated correctly")


def test_all_resource_types_have_sample_data():
    """Test that all resource types have sample data."""
    service = ExcelGeneratorService()
    template_bytes = service.generate_template(TemplateType.FULL)
    wb = load_workbook(io.BytesIO(template_bytes))

    resource_sheets = [
        "AWS_EC2",
        "AWS_VPC",
        "AWS_Subnet",
        "AWS_SecurityGroup",
        "AWS_S3",
        "AWS_RDS",
        "Azure_VM",
        "Azure_VNet",
        "Azure_Subnet",
        "Azure_NSG",
        "Azure_Storage",
        "Azure_SQL",
    ]

    for sheet_name in resource_sheets:
        sheet = wb[sheet_name]
        resource_name_sample = sheet.cell(2, 1).value
        assert resource_name_sample is not None and resource_name_sample != "", (
            f"{sheet_name} should have sample ResourceName in row 2"
        )

    print("PASS: All resource types have sample data")


def run_all_tests():
    """Run all tests."""
    print("=" * 80)
    print("RUNNING EXCEL TEMPLATE IMPROVEMENT TESTS")
    print("=" * 80)
    print()

    try:
        test_required_field_markers()
        test_sample_data_row()
        test_parser_strips_asterisks()
        test_readme_instructions()
        test_all_resource_types_have_sample_data()

        print()
        print("=" * 80)
        print("ALL TESTS PASSED!")
        print("=" * 80)
        print()
        print("Summary of improvements:")
        print("  1. Required fields marked with asterisk (*)")
        print("  2. Required fields have darker purple background (#7D00C7)")
        print("  3. Optional fields have standard purple background (#A100FF)")
        print("  4. Sample data added to row 2 of each sheet")
        print("  5. Parser correctly strips asterisks from headers")
        print("  6. README updated with detailed instructions")

    except AssertionError as e:
        print()
        print("=" * 80)
        print("TEST FAILED!")
        print("=" * 80)
        print(f"Error: {e}")
        raise


if __name__ == "__main__":
    run_all_tests()
