"""End-to-end test: Excel upload -> Compliance check"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.services.excel_parser import ExcelParserService
from app.core.database import SessionLocal
from app.agents.nodes import AgentNodes
from openpyxl import Workbook
from io import BytesIO

print("=" * 80)
print("END-TO-END TEST: Excel Upload -> Compliance Check")
print("=" * 80)

# Create test Excel file (same as user's example)
wb = Workbook()
ws = wb.active
ws.title = "Azure_VM"

headers = [
    "ResourceName*",
    "Environment*",
    "Project*",
    "Owner",
    "CostCenter",
    "Tags",
    "ResourceGroup*",
    "Location*",
    "VMSize*",
    "OSType*",
    "ImagePublisher",
    "ImageOffer",
    "ImageSKU",
    "ImageVersion",
]
for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=header)

# Test 1: Resource WITH Project tag (should PASS)
print("\nTEST 1: Resource WITH Project='abc' (should PASS)")
print("-" * 80)

data_row = [
    "web-vm-01",
    "Production",
    "abc",
    "john.doe@example.com",
    "IT-1234",
    '{"Application": "WebServer"}',
    "rg-myproject-prod",
    "eastus",
    "Standard_D2s_v3",
    "Linux",
    "Canonical",
    "UbuntuServer",
    "18.04-LTS",
    "latest",
]
for col_idx, value in enumerate(data_row, start=1):
    ws.cell(row=2, column=col_idx, value=value)

excel_bytes = BytesIO()
wb.save(excel_bytes)
excel_bytes.seek(0)

# Parse Excel
parser = ExcelParserService()
parse_result = parser.parse_excel_file(excel_bytes.read())

print(f"Parse Success: {parse_result.success}")
print(f"Resources: {parse_result.resource_count}")

if parse_result.resources:
    resource = parse_result.resources[0]
    print(f"Resource Tags: {resource.properties.get('Tags', {})}")

    # Run compliance check
    db = SessionLocal()
    state = {
        "session_id": "e2e-test-001",
        "resources": [resource.model_dump()],
        "messages": [],
        "workflow_state": "initial",
    }

    nodes = AgentNodes(db)
    result = nodes.compliance_checker(state)

    print(
        f"\nCompliance Result: {'PASSED' if result['compliance_passed'] else 'FAILED'}"
    )
    print(f"Violations: {len(result.get('compliance_violations', []))}")

    if result["compliance_passed"]:
        print("[SUCCESS] Test 1 PASSED: Resource with Project tag passed compliance")
    else:
        print("[FAILED] Test 1 FAILED: Should have passed")
        for v in result.get("compliance_violations", []):
            print(f"  - {v['resource']}: {v['issue']}")
        sys.exit(1)

    db.close()

# Test 2: Resource WITHOUT Project tag (should FAIL)
print("\n" + "=" * 80)
print("TEST 2: Resource WITHOUT Project tag (should FAIL)")
print("-" * 80)

wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Azure_VM"

for col_idx, header in enumerate(headers, start=1):
    ws2.cell(row=1, column=col_idx, value=header)

# Remove Project column value
data_row2 = [
    "web-vm-02",
    "Production",
    "",  # Empty Project!
    "john.doe@example.com",
    "IT-1234",
    '{"Application": "WebServer"}',
    "rg-myproject-prod",
    "eastus",
    "Standard_D2s_v3",
    "Linux",
    "Canonical",
    "UbuntuServer",
    "18.04-LTS",
    "latest",
]
for col_idx, value in enumerate(data_row2, start=1):
    ws2.cell(row=2, column=col_idx, value=value)

excel_bytes2 = BytesIO()
wb2.save(excel_bytes2)
excel_bytes2.seek(0)

parser2 = ExcelParserService()
parse_result2 = parser2.parse_excel_file(excel_bytes2.read())

print(f"Parse Success: {parse_result2.success}")
print(f"Resources: {parse_result2.resource_count}")

if parse_result2.resources:
    resource2 = parse_result2.resources[0]
    print(f"Resource Tags: {resource2.properties.get('Tags', {})}")

    db2 = SessionLocal()
    state2 = {
        "session_id": "e2e-test-002",
        "resources": [resource2.model_dump()],
        "messages": [],
        "workflow_state": "initial",
    }

    nodes2 = AgentNodes(db2)
    result2 = nodes2.compliance_checker(state2)

    print(
        f"\nCompliance Result: {'PASSED' if result2['compliance_passed'] else 'FAILED'}"
    )
    print(f"Violations: {len(result2.get('compliance_violations', []))}")

    if not result2["compliance_passed"]:
        print("[SUCCESS] Test 2 PASSED: Resource without Project tag failed compliance")
        for v in result2.get("compliance_violations", []):
            print(f"  - {v['resource']}: {v['issue']}")
    else:
        print("[FAILED] Test 2 FAILED: Should have failed compliance")
        sys.exit(1)

    db2.close()

print("\n" + "=" * 80)
print("ALL END-TO-END TESTS PASSED!")
print("Excel metadata columns are correctly merged into Tags")
print("Compliance checker correctly validates Project tag")
print("=" * 80)
