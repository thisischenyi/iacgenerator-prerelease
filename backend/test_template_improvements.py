"""Test script to verify Excel template improvements."""

import io
from openpyxl import load_workbook
from app.services.excel_generator import ExcelGeneratorService
from app.schemas import TemplateType


def test_template_improvements():
    """Test that templates have required field indicators and sample data."""
    service = ExcelGeneratorService()

    # Generate a full template
    template_bytes = service.generate_template(TemplateType.FULL)

    # Load it back with openpyxl
    wb = load_workbook(io.BytesIO(template_bytes))

    print("=" * 80)
    print("EXCEL TEMPLATE VERIFICATION")
    print("=" * 80)

    # Test AWS_EC2 sheet
    print("\n[AWS_EC2 Sheet]")
    sheet = wb["AWS_EC2"]

    # Check headers (row 1)
    print("\nHeaders (Row 1):")
    headers = []
    for col in range(1, 21):  # Check first 20 columns
        cell = sheet.cell(1, col)
        if cell.value:
            headers.append(cell.value)
            has_asterisk = "*" in str(cell.value)
            bg_color = cell.fill.start_color.rgb if cell.fill.start_color else "None"
            print(f"  {cell.value:30} | Required: {has_asterisk:5} | BG: {bg_color}")

    # Check sample data (row 2)
    print("\nSample Data (Row 2):")
    for col_idx, header in enumerate(headers[:10], start=1):  # Show first 10
        cell = sheet.cell(2, col_idx)
        value = cell.value if cell.value else "(empty)"
        print(f"  {header:30} = {value}")

    # Test Azure_VM sheet
    print("\n" + "=" * 80)
    print("\n[Azure_VM Sheet]")
    sheet = wb["Azure_VM"]

    # Check headers
    print("\nHeaders (Row 1):")
    headers = []
    for col in range(1, 21):
        cell = sheet.cell(1, col)
        if cell.value:
            headers.append(cell.value)
            has_asterisk = "*" in str(cell.value)
            bg_color = cell.fill.start_color.rgb if cell.fill.start_color else "None"
            print(f"  {cell.value:30} | Required: {has_asterisk:5} | BG: {bg_color}")

    # Check sample data
    print("\nSample Data (Row 2):")
    for col_idx, header in enumerate(headers[:10], start=1):
        cell = sheet.cell(2, col_idx)
        value = cell.value if cell.value else "(empty)"
        print(f"  {header:30} = {value}")

    # Test README sheet
    print("\n" + "=" * 80)
    print("\n[README Sheet]")
    readme = wb["README"]
    print("\nInstructions:")
    for row in range(2, 9):  # Rows 2-8 contain instructions
        topic = readme.cell(row, 1).value
        detail = readme.cell(row, 2).value
        print(f"\n{topic}:")
        print(f"  {detail}")

    print("\n" + "=" * 80)
    print("\n✅ VERIFICATION COMPLETE")
    print("=" * 80)

    # Verify key requirements
    assert "ResourceName*" in [
        sheet.cell(1, c).value for c in range(1, 30) for sheet in [wb["AWS_EC2"]]
    ], "AWS_EC2 should have ResourceName* with asterisk"

    assert wb["AWS_EC2"].cell(2, 1).value == "web-server-01", (
        "AWS_EC2 sample data should have ResourceName = 'web-server-01'"
    )

    assert "ResourceName*" in [wb["Azure_VM"].cell(1, c).value for c in range(1, 30)], (
        "Azure_VM should have ResourceName* with asterisk"
    )

    print("\n✅ All assertions passed!")


if __name__ == "__main__":
    test_template_improvements()
